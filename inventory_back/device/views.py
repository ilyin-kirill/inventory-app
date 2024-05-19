import os

from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets
from django.http import FileResponse, HttpResponse

from device.models import Device
from device.serializers import DeviceSerializer

from openpyxl import load_workbook

class DeviceViewSet(viewsets.ModelViewSet):
    serializer_class = DeviceSerializer
    queryset = Device.objects.all()



class FileProperty:
    start_row=4
    start_col=1
    final_row=17
    final_col=23

    def check_row(self)->bool:
        if self.start_row>self.final_row:
            return False
        return True
def get_file(request):
    wb = load_workbook(os.getcwd()+'/device/file/inv-3-blank.xlsx')
    ws=wb["стр2"]
    FP = FileProperty()
    for index, device in enumerate(Device.objects.all()):
        ws.cell(row=FP.start_row, column=1, value=index)
        ws.cell(row=FP.start_row, column=3, value=device.account)
        ws.cell(row=FP.start_row, column=6, value=device.characteristics)
        ws.cell(row=FP.start_row, column=7, value=device.code)
        ws.cell(row=FP.start_row, column=9, value=device.code_okei)
        ws.cell(row=FP.start_row, column=10, value=device.count_name)
        ws.cell(row=FP.start_row, column=12, value=device.price)
        ws.cell(row=FP.start_row, column=14, value=device.inventory_num)
        ws.cell(row=FP.start_row, column=15, value=device.pasport)
        ws.cell(row=FP.start_row, column=17, value=device.fact_count)
        ws.cell(row=FP.start_row, column=18, value=device.fact_sum)
        ws.cell(row=FP.start_row, column=20, value=device.accounting_count)
        ws.cell(row=FP.start_row, column=21, value=device.accounting_sum)
        if FP.check_row:
            FP.start_row += 1
        else:
            break
    wb.save(os.getcwd()+'/device/file/result.xlsx')
    return FileResponse(open(os.getcwd()+'/device/file/result.xlsx', 'rb'))